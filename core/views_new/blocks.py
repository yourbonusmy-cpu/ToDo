from datetime import datetime

from django.contrib.auth.hashers import check_password
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from core.models import (
    GroupTemplate,
    PeriodType,
    ScheduleType,
    BlockWeather,
    DeletedObject,
)
from django.shortcuts import get_object_or_404, redirect
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os
import zipfile
from io import BytesIO
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.db.models import Prefetch
from core.models import Block, BlockTask, TaskTemplate
import json
from django.contrib.auth.decorators import login_required
from core.crypto import encrypt_text
from core.pagination import BlockCursorPagination
from core.serializers_ import BlockSerializer
from django.db.models import Q, Count
from rest_framework.generics import ListAPIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated


@login_required
def block_create(request, block_id=None):
    # ==========================
    # LOAD task_templates
    # ==========================
    task_templates = TaskTemplate.objects.filter(owner=request.user)

    period_type = request.GET.get("period_type", PeriodType.NONE)
    schedule_type = request.GET.get("schedule_type", ScheduleType.NONE)

    if period_type != PeriodType.NONE:
        task_templates = task_templates.filter(period_type=period_type)
    if schedule_type != ScheduleType.NONE:
        task_templates = task_templates.filter(schedule_type=schedule_type)

    sorted_task_templates = task_templates.annotate(
        is_floating=(
            Q(schedule_type=ScheduleType.FLOATING)
            & Q(period_type__in=[PeriodType.WEEK, PeriodType.MONTH, PeriodType.YEAR])
        ),
        is_fixed=(
            Q(schedule_type=ScheduleType.FIXED)
            & Q(period_type__in=[PeriodType.WEEK, PeriodType.MONTH, PeriodType.YEAR])
        ),
    ).order_by(
        "-is_floating",
        "-fixed_weekday",
        "-fixed_day_of_month",
        "-fixed_month_of_year",
        "-next_available_at",
        "-selected_count",
        "-priority",
        "-last_used_at",
    )

    groups_templates = GroupTemplate.objects.filter(
        owner=request.user
    ).prefetch_related(
        Prefetch("tasks", queryset=TaskTemplate.objects.only("id", "icon"))
    )

    # ==========================
    # LOAD BLOCK (edit mode)
    # ==========================
    block = None
    block_data = None

    if block_id:
        block = get_object_or_404(Block, id=block_id, owner=request.user)
        block_tasks = block.block_tasks.order_by("position")

        block_data = {
            "id": block.id,
            "title": block.title,
            "weather": (
                getattr(block.weather, "data", None)
                if hasattr(block, "weather")
                else None
            ),
            "weather_city": (
                getattr(block.weather, "city", None)
                if hasattr(block, "weather")
                else None
            ),
            "target_date": block.target_date.isoformat() if block.target_date else None,
            "block_tasks": [
                {
                    "id": bt.id,
                    "title": bt.title,
                    "description": bt.description,
                    "amount": bt.amount,
                    "time": bt.time,
                    "icon": str(bt.icon) if bt.icon else "",
                    "is_encrypted": bt.is_encrypted,
                    "is_hidden": bt.is_hidden,
                    "template_id": bt.template.id if bt.template else None,
                }
                for bt in block_tasks
            ],
        }

    # ==========================
    # POST HANDLER
    # ==========================
    if request.method == "POST":

        # ---------- JSON ----------
        if request.content_type == "application/json":
            body = json.loads(request.body)

            title = body.get("title", "Без названия")
            target_date = body.get("target_date")
            block_tasks_list = body.get("block_tasks", [])

            with_weather = body.get("with_weather", False)
            city = body.get("weather_city", "Москва")
            weather_data_raw = body.get("weather_data")

        # ---------- FORM ----------
        else:
            title = request.POST.get("title", "Без названия")
            target_date = request.POST.get("target_date") or None

            block_tasks_json = request.POST.get("block_tasks", "[]")
            block_tasks_list = json.loads(block_tasks_json)

            with_weather = request.POST.get("with_weather") == "true"
            city = request.POST.get("weather_city", "Москва")
            weather_data_raw = request.POST.get("weather_data")

        # parse date safely
        if target_date:
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()

        # ==========================
        # UPDATE BLOCK
        # ==========================
        if block:
            block.title = title
            block.target_date = target_date
            block.save()

            # ---------- WEATHER ----------
            if with_weather and weather_data_raw:
                try:
                    BlockWeather.objects.update_or_create(
                        block=block,
                        defaults={
                            "city": city,
                            "data": json.loads(weather_data_raw),
                        },
                    )
                except json.JSONDecodeError:
                    pass  # можно логировать

            elif not with_weather:
                BlockWeather.objects.filter(block=block).delete()

            # ---------- TASKS ----------
            existing_block_tasks = {t.id: t for t in block.block_tasks.all()}
            incoming_ids = set()

            for idx, task_data in enumerate(block_tasks_list):

                template_id = task_data.get("template_id")
                if not template_id:
                    raise ValueError("Каждая задача должна иметь template_id")

                template = get_object_or_404(
                    TaskTemplate, id=template_id, owner=request.user
                )

                task_id = task_data.get("id")
                description = task_data.get("description", "")
                is_encrypted = task_data.get("is_encrypted", False)
                is_hidden = task_data.get("is_hidden", False)

                # encryption
                if is_encrypted:
                    password = task_data.get("password")
                    decrypted = task_data.get("decrypted")

                    if decrypted and password:
                        description = encrypt_text(description, password)
                    elif task_id in existing_block_tasks:
                        description = existing_block_tasks[task_id].description

                if task_id and task_id in existing_block_tasks:
                    bt = existing_block_tasks[task_id]
                    bt.title = task_data.get("title", bt.title)
                    bt.description = description
                    bt.amount = task_data.get("amount", bt.amount)
                    bt.time = task_data.get("time", bt.time)
                    bt.icon = task_data.get("icon", bt.icon)
                    bt.is_encrypted = is_encrypted
                    bt.is_hidden = is_hidden
                    bt.template = template
                    bt.position = idx
                    bt.save()
                else:
                    bt = BlockTask.objects.create(
                        block=block,
                        template=template,
                        title=task_data.get("title"),
                        description=description,
                        amount=task_data.get("amount", 0),
                        time=task_data.get("time", 0),
                        icon=task_data.get("icon", ""),
                        is_encrypted=is_encrypted,
                        is_hidden=is_hidden,
                        position=idx,
                    )

                incoming_ids.add(bt.id)

            # delete removed tasks
            # BlockTask.objects.filter(
            #     id__in=set(existing_tasks.keys()) - incoming_ids
            # ).delete()
            removed_ids = set(existing_block_tasks.keys()) - incoming_ids

            removed_block_tasks = BlockTask.objects.filter(id__in=removed_ids)

            DeletedObject.objects.bulk_create(
                [
                    DeletedObject(
                        obj_uuid=t.uuid,
                        object_type=DeletedObject.ObjectType.BLOCKTASK,
                        user=request.user,
                        device_uuid=getattr(request, "device_uuid", None),
                    )
                    for t in removed_block_tasks
                ]
            )

            removed_block_tasks.delete()

        # ==========================
        # CREATE BLOCK
        # ==========================
        else:
            block = Block.objects.create(
                title=title,
                owner=request.user,
                target_date=target_date,
            )

            for idx, task_data in enumerate(block_tasks_list):

                template_id = task_data.get("template_id")
                if not template_id:
                    raise ValueError("Каждая задача должна иметь template_id")

                template = get_object_or_404(
                    TaskTemplate, id=template_id, owner=request.user
                )

                description = task_data.get("description", "")
                is_encrypted = task_data.get("is_encrypted", False)
                is_hidden = task_data.get("is_hidden", False)

                if is_encrypted:
                    password = task_data.get("password")
                    if password:
                        description = encrypt_text(description, password)

                BlockTask.objects.create(
                    block=block,
                    template=template,
                    title=task_data.get("title"),
                    description=description,
                    amount=task_data.get("amount", 0),
                    time=task_data.get("time", 0),
                    icon=task_data.get("icon", ""),
                    is_encrypted=is_encrypted,
                    is_hidden=is_hidden,
                    position=idx,
                )

            # ---------- WEATHER ----------
            if with_weather and weather_data_raw:
                try:
                    BlockWeather.objects.create(
                        block=block,
                        city=city,
                        data=json.loads(weather_data_raw),
                    )
                except json.JSONDecodeError:
                    pass

        return redirect("/")

    # ==========================
    # RENDER
    # ==========================
    return render(
        request,
        "core/blocks/block_create.html",
        {
            "task_templates": sorted_task_templates,
            "group_templates": groups_templates,
            "block_json": json.dumps(block_data) if block_data else "null",
            "MEDIA_URL": settings.MEDIA_URL,
        },
    )


class BlockListView(ListAPIView):
    serializer_class = BlockSerializer
    pagination_class = BlockCursorPagination
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        request = self.request

        q = request.GET.get("q")
        block_task_ids = request.GET.getlist("block_tasks")
        weekdays = request.GET.getlist("weekdays")

        queryset = Block.objects.filter(owner=request.user, is_hidden=False)

        # -------------------
        # SEARCH
        # -------------------
        if q:
            queryset = queryset.filter(title__icontains=q)

        # -------------------
        # TASK FILTER
        # -------------------
        if block_task_ids:
            try:
                block_task_ids = list(map(int, block_task_ids))
            except ValueError:
                return Block.objects.none()

            queryset = queryset.annotate(
                matched_block_tasks=Count(
                    "block_tasks",
                    filter=Q(block_tasks__template__id__in=block_task_ids),
                    distinct=True,
                )
            ).filter(matched_block_tasks=len(block_task_ids))

        # -------------------
        # WEEKDAY FILTER (NEW)
        # -------------------
        if weekdays:
            try:
                weekdays = list(map(int, weekdays))
            except ValueError:
                return Block.objects.none()

            q_filter = Q()

            for day in weekdays:
                q_filter |= Q(target_date__week_day=((day + 2) % 7))

            queryset = queryset.filter(q_filter)

        return queryset.prefetch_related("block_tasks__template").order_by(
            "-created_at", "-id"
        )


@login_required
def block_detail(request, block_id):
    block = get_object_or_404(Block, id=block_id, owner=request.user)
    block_tasks = block.block_tasks.order_by("position")
    return render(
        request,
        "core/blocks/block_detail.html",
        {"m_block": block, "block_tasks": block_tasks},
    )


@login_required
@csrf_exempt
def hide_block(request, block_id):
    """
    Переключает состояние is_hidden у блока
    """
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Only POST allowed"}, status=405
        )

    block = get_object_or_404(Block, id=block_id, owner=request.user)
    block.is_hidden = not block.is_hidden
    block.save()
    return JsonResponse({"status": "ok", "is_hidden": block.is_hidden})


@require_POST
@login_required
def delete_block(request, block_id):
    block = get_object_or_404(Block, id=block_id, owner=request.user)

    password = request.POST.get("password")
    if not password:
        return JsonResponse(
            {"status": "error", "message": "Password required"}, status=400
        )

    if not check_password(password, request.user.password):
        return JsonResponse(
            {"status": "error", "message": "Wrong password"}, status=403
        )

    device_uuid = getattr(request, "device_uuid", None)

    with transaction.atomic():
        DeletedObject.objects.create(
            obj_uuid=block.uuid,
            object_type=DeletedObject.ObjectType.BLOCK,
            user=request.user,
            device_uuid=device_uuid,
        )

        tasks = BlockTask.objects.filter(block=block)

        DeletedObject.objects.bulk_create(
            [
                DeletedObject(
                    obj_uuid=t.uuid,
                    object_type=DeletedObject.ObjectType.BLOCKTASK,
                    user=request.user,
                    device_uuid=device_uuid,
                )
                for t in tasks
            ]
        )

        block.delete()

    return JsonResponse({"status": "ok"})


def download_blocks_xlsx(request):
    blocks = Block.objects.filter(owner=request.user).prefetch_related("block_tasks")

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocks"

    row = 1

    for block in blocks:
        # Заголовок блока — жирный, объединяем колонки
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(
            row=row,
            column=1,
            value=f"Block: {block.title}",
        )
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="left")
        row += 2

        # Заголовки задач
        headers = ["Task Title", "Icon", "Description", "Amount", "Time"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for task in block.block_tasks.all():
            # Название задачи
            ws.cell(row=row, column=1, value=task.title)
            ws.cell(row=row, column=3, value=task.description)
            ws.cell(row=row, column=4, value=task.amount)
            ws.cell(row=row, column=5, value=task.time)

            # Иконка
            if task.icon:
                try:
                    icon_str = str(task.icon)
                    if icon_str.startswith("http"):
                        icon_str = icon_str.split("/media/")[-1]
                    icon_path = os.path.join(settings.MEDIA_ROOT, icon_str)

                    if os.path.exists(icon_path):
                        img = Image(icon_path)
                        img.width = 56
                        img.height = 56
                        img.anchor = f"B{row}"
                        ws.add_image(img)

                except Exception as e:
                    print("Image error:", e)

            # Высота строки под иконку
            ws.row_dimensions[row].height = 56

            row += 1

        row += 2  # пустая строка между блоками

    # Ширина колонок для лучшего отображения
    ws.column_dimensions["A"].width = 30  # Task Title
    ws.column_dimensions["B"].width = 12  # Icon
    ws.column_dimensions["C"].width = 40  # Description
    ws.column_dimensions["D"].width = 10  # Amount
    ws.column_dimensions["E"].width = 10  # Time

    # Сохраняем файл
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    response = HttpResponse(
        file_buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="blocks.xlsx"'
    return response


def download_blocks_json_zip(request):

    blocks = Block.objects.filter(owner=request.user).prefetch_related("block_tasks")

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        blocks_json = []
        added_icons = set()

        for block in blocks:

            block_data = {
                "title": block.title,
                "priority": block.priority,
                "block_tasks": [],
            }

            for task in block.block_tasks.all():

                icon_archive_path = ""

                if task.icon:

                    icon_str = str(task.icon)

                    # убираем домен
                    if icon_str.startswith("http"):
                        icon_str = icon_str.split("/media/")[-1]

                    icon_path = os.path.join(settings.MEDIA_ROOT, icon_str)

                    print("ICON:", icon_str)
                    print("PATH:", icon_path)

                    if os.path.exists(icon_path):

                        icon_name = os.path.basename(icon_str)
                        icon_archive_path = f"icons/{icon_name}"

                        if icon_archive_path not in added_icons:
                            zip_file.write(icon_path, icon_archive_path)
                            added_icons.add(icon_archive_path)

                task_data = {
                    "title": task.title,
                    "description": task.description,
                    "position": task.position,
                    "amount": task.amount,
                    "time": task.time,
                    "icon": icon_archive_path,
                }

                block_data["block_tasks"].append(task_data)

            blocks_json.append(block_data)

        zip_file.writestr(
            "blocks.json",
            json.dumps(blocks_json, ensure_ascii=False, indent=2),
        )

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="blocks_backup.zip"'

    return response
