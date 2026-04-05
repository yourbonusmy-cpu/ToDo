from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Prefetch
from core.models import Block, BlockTask
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os
import json
import zipfile
from io import BytesIO

from django.conf import settings
from django.http import HttpResponse


from django.db.models import Prefetch
from core.models import Block, BlockTask, TaskTemplate  # 👈 добавить


def home(request):
    if not request.user.is_authenticated:
        return render(request, "core/landing.html")

    blocks_qs = (
        Block.objects.filter(owner=request.user, is_hidden=False)
        .prefetch_related(
            Prefetch(
                "tasks",
                queryset=BlockTask.objects.order_by("position"),
            )
        )
        .order_by("-created_at")
    )

    paginator = Paginator(blocks_qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ✅ Только пользовательские шаблоны
    system_tasks = TaskTemplate.objects.filter(owner=request.user).order_by("title")

    return render(
        request,
        "core/dashboard.html",
        {
            "page_obj": page_obj,
            "blocks": page_obj.object_list,
            "system_tasks": system_tasks,
        },
    )


@login_required
def block_detail(request, block_id):
    block = get_object_or_404(Block, id=block_id, owner=request.user)
    block_tasks = block.tasks.order_by("position")
    return render(
        request,
        "core/block_detail.html",
        {"m_block": block, "block_tasks": block_tasks},
    )


def download_blocks_xlsx(request):
    blocks = Block.objects.filter(owner=request.user).prefetch_related("tasks")

    wb = Workbook()
    ws = wb.active
    ws.title = "Blocks"

    row = 1

    for block in blocks:
        # Заголовок блока — жирный, объединяем колонки
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(
            row=row,
            column=1,
            value=f"Block: {block.title}",
        )
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="left")
        row += 2

        # Заголовки задач
        headers = ["Task Title", "Icon", "Description", "Amount", "Time"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for task in block.tasks.all():
            # Название задачи
            ws.cell(row=row, column=1, value=task.title)
            ws.cell(row=row, column=3, value=task.description)
            ws.cell(row=row, column=4, value=task.amount)
            ws.cell(row=row, column=5, value=task.time)

            # Иконка
            if task.icon:
                try:
                    icon_str = str(task.icon)
                    if icon_str.startswith("http"):
                        icon_str = icon_str.split("/media/")[-1]
                    icon_path = os.path.join(settings.MEDIA_ROOT, icon_str)

                    if os.path.exists(icon_path):
                        img = Image(icon_path)
                        img.width = 56
                        img.height = 56
                        img.anchor = f"B{row}"  # Иконка в колонке B
                        ws.add_image(img)

                except Exception as e:
                    print("Image error:", e)

            # Высота строки под иконку
            ws.row_dimensions[row].height = 56

            row += 1

        row += 2  # пустая строка между блоками

    # Ширина колонок для лучшего отображения
    ws.column_dimensions["A"].width = 30  # Task Title
    ws.column_dimensions["B"].width = 12  # Icon
    ws.column_dimensions["C"].width = 40  # Description
    ws.column_dimensions["D"].width = 10  # Amount
    ws.column_dimensions["E"].width = 10  # Time

    # Сохраняем файл
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    response = HttpResponse(
        file_buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="blocks.xlsx"'
    return response


def download_blocks_json_zip(request):

    blocks = Block.objects.filter(owner=request.user).prefetch_related("tasks")

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        blocks_json = []
        added_icons = set()

        for block in blocks:

            block_data = {
                "title": block.title,
                "priority": block.priority,
                "tasks": [],
            }

            for task in block.tasks.all():

                icon_archive_path = ""

                if task.icon:

                    icon_str = str(task.icon)

                    # убираем домен
                    if icon_str.startswith("http"):
                        icon_str = icon_str.split("/media/")[-1]

                    icon_path = os.path.join(settings.MEDIA_ROOT, icon_str)

                    print("ICON:", icon_str)
                    print("PATH:", icon_path)

                    if os.path.exists(icon_path):

                        icon_name = os.path.basename(icon_str)
                        icon_archive_path = f"icons/{icon_name}"

                        if icon_archive_path not in added_icons:
                            zip_file.write(icon_path, icon_archive_path)
                            added_icons.add(icon_archive_path)

                task_data = {
                    "title": task.title,
                    "description": task.description,
                    "position": task.position,
                    "amount": task.amount,
                    "time": task.time,
                    "icon": icon_archive_path,
                }

                block_data["tasks"].append(task_data)

            blocks_json.append(block_data)

        zip_file.writestr(
            "blocks.json",
            json.dumps(blocks_json, ensure_ascii=False, indent=2),
        )

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="blocks_backup.zip"'

    return response
